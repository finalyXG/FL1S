from tqdm import tqdm
import datetime
import io
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from IPython import display
import os
import tensorflow as tf
import seaborn as sns
from sklearn.manifold import TSNE
from absl import app, flags
from tensorboard.plugins.hparams import api as hp
import shutil
import time
import random
import pickle
import copy
import openpyxl
class Trainer:
    def __init__(self, client_name, version_num, client_data,all_test_x,all_test_y, initial_feature_center, pre_features_central, cls, teacher_list, discriminator, generator, config, hparams):
        self.client_name = client_name
        self.version_num = str(version_num)
        self.cls = cls
        self.teacher_list = teacher_list
        self.discriminator = discriminator
        self.generator = generator
        self.config = config
        self.initial_client_ouput_feat_epochs = config.initial_client_ouput_feat_epochs
        tf.random.set_seed(config.random_seed)
        np.random.seed(config.random_seed)
        random.seed(config.random_seed)
        self.pre_features_central = pre_features_central
        self.initial_feature_center = initial_feature_center
        #split data
        (train_data, test_data) = client_data
        self.all_test_x,self.all_test_y = all_test_x,all_test_y
        self.train_x,self.train_y = zip(*train_data)
        self.train_data_num = len(self.train_y)
        self.test_x,self.test_y = zip(*test_data)
        self.class_rate = self.train_y.count(0)/self.train_y.count(1)
        print("class_rate",self.class_rate)
        self.train_x,self.train_y = np.array(self.train_x),np.array(self.train_y)
        self.test_x,self.test_y = np.array(self.test_x),np.array(self.test_y)
        ## show the distribution of label
        # for element in set(self.train_y):
        #     print(element," count: ", list(self.train_y).count(element))

        self.train_data = tf.data.Dataset.from_tensor_slices(
        (self.train_x,self.train_y)).shuffle(len(self.train_y))
        self.test_data = tf.data.Dataset.from_tensor_slices(
        (self.test_x,self.test_y)).shuffle(len(self.test_y)).batch(hparams['batch_size'],drop_remainder=True)
        self.all_test_data = tf.data.Dataset.from_tensor_slices(
        (all_test_x,all_test_y)).shuffle(len(all_test_y)).batch(hparams['batch_size'],drop_remainder=True)
        
        self.local_cls_acc_list = []  
        self.global_cls_acc_list = []
        self.global_cls_f1_list = []
        self.global_cls_recall_list = []
        self.global_cls_precision_list = []
        self.GAN_version = "ACGAN" #hparams['gan_version']
        self.batch_size = hparams['batch_size']
        self.learning_rate = hparams['learning_rate']

        self.cos_loss_weight = hparams['cos_loss_weight']
        self.original_cls_loss_weight = hparams['original_cls_loss_weight']
        self.feat_loss_weight = hparams['feat_loss_weight']

        self.gp_weight = config.gp_weight
        self.num_classes = config.num_classes
        self.discriminator_extra_steps = config.discriminator_extra_steps
        self.img_save_path = "./generate_img"
        self.latent_dim = config.latent_dim

        self.cls_optimizer = tf.keras.optimizers.legacy.Adam(self.learning_rate)
        self.disc_optimizer = tf.keras.optimizers.legacy.Adam(self.learning_rate)
        self.gen_optimizer = tf.keras.optimizers.legacy.Adam(self.learning_rate)
        
        self.loss_fn_gan_binary = tf.keras.losses.BinaryCrossentropy(from_logits=True)
        self.loss_fn_gan_categorical = tf.keras.losses.CategoricalCrossentropy(from_logits=True) # output not normalized
        self.disc_test_binary_accuracy = tf.keras.metrics.BinaryAccuracy(name='disc_test_binary_accuracy')
        self.disc_test_categorical_accuracy = tf.keras.metrics.CategoricalAccuracy(name='disc_test_categorical_accuracy')
        self.disc_train_binary_accuracy = tf.keras.metrics.BinaryAccuracy(name='disc_train_binary_accuracy')
        self.disc_train_categorical_accuracy = tf.keras.metrics.CategoricalAccuracy(name='disc_train_categorical_accuracy')

        self.cls_train_loss = tf.keras.metrics.Mean(name='cls_train_loss')
        self.cls_train_distance_loss = tf.keras.metrics.Mean(name='cls_train_distance_loss')
        self.cls_train_classify_loss = tf.keras.metrics.Mean(name='cls_train_classify_loss')
        self.cls_train_feature_loss = tf.keras.metrics.Mean(name='cls_train_feature_loss')
        self.cls_test_loss = tf.keras.metrics.Mean(name='cls_test_loss')
    
        if config.dataset != "elliptic":
            self.img_loss_fn_cls = tf.keras.losses.SparseCategoricalCrossentropy(from_logits=False)
            self.feature_loss_fn_cls = tf.keras.losses.SparseCategoricalCrossentropy(from_logits=False)
            self.cls_train_accuracy = tf.keras.metrics.SparseCategoricalAccuracy(name='cls_train_accuracy')
            self.cls_test_accuracy = tf.keras.metrics.SparseCategoricalAccuracy(name='cls_test_accuracy')
            self.cls_global_test_accuracy = tf.keras.metrics.SparseCategoricalAccuracy(name='cls_global_test_accuracy')
        else:
            # class_weights = {0: 0.3, 1: 0.7}
            self.config.importance_rate = (config.train_data_importance_rate * self.class_rate).astype('float32')
            print("importance_rate",self.config.importance_rate)
            self.weights = tf.constant(self.config.importance_rate)
            self.cls_train_accuracy = tf.keras.metrics.BinaryAccuracy(name='cls_train_accuracy')
            self.cls_test_accuracy = tf.keras.metrics.BinaryAccuracy(name='cls_test_accuracy')
            self.cls_global_test_accuracy = tf.keras.metrics.BinaryAccuracy(name='cls_global_test_accuracy')

        self.cls_train_elliptic_recall = tf.keras.metrics.Recall(name='elliptic_train_recall')
        self.cls_train_elliptic_precision = tf.keras.metrics.Precision(name='elliptic_train_precision')
        self.cls_train_elliptic_f1 = tf.keras.metrics.F1Score(threshold=0.5, name='elliptic_train_f1score')

        self.cls_test_elliptic_recall = tf.keras.metrics.Recall(name='elliptic_test_recall')
        self.cls_test_elliptic_precision = tf.keras.metrics.Precision(name='elliptic_test_precision')
        self.cls_test_elliptic_f1 = tf.keras.metrics.F1Score(threshold=0.5, name='elliptic_test_f1score')

        self.cls_global_test_elliptic_recall = tf.keras.metrics.Recall(name='elliptic_global_test_recall')
        self.cls_global_test_elliptic_precision = tf.keras.metrics.Precision(name='elliptic_global_test_precision')
        self.cls_global_test_elliptic_f1 = tf.keras.metrics.F1Score(threshold=0.5,name='elliptic_global_test_f1score')  

        self.disc_train_loss = tf.keras.metrics.Mean(name='disc_train_loss')
        self.gen_train_loss = tf.keras.metrics.Mean(name='gen_train_loss')  
        self.disc_test_loss = tf.keras.metrics.Mean(name='disc_test_loss')
        self.gen_test_loss = tf.keras.metrics.Mean(name='gen_test_loss')

        current_time = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")

        GAN_train_log_dir = 'logs/GAN_gradient_tape/' +client_name +"/"+ self.version_num + '/train'
        GAN_test_log_dir = 'logs/GAN_gradient_tape/' +client_name +"/"+ self.version_num + '/test'
        CLS_train_log_dir = 'logs/CLS_gradient_tape/' +client_name +"/"+ self.version_num + '/train'
        CLS_test_log_dir = 'logs/CLS_gradient_tape/' +client_name +"/"+ self.version_num + '/test'
        CLS_compare_test_log_dir = 'logs/CLS_gradient_tape/' +client_name +"/"+ self.version_num + '/global'

        # generator_img_logdir = "logs/train_data/" +client_name +"/"+ self.version_num
        # self.GAN_train_summary_writer = tf.summary.create_file_writer(GAN_train_log_dir)
        # self.GAN_test_summary_writer = tf.summary.create_file_writer(GAN_test_log_dir)
        self.CLS_train_summary_writer = tf.summary.create_file_writer(CLS_train_log_dir)
        self.CLS_test_summary_writer = tf.summary.create_file_writer(CLS_test_log_dir)
        self.CLS_compare_test_acc_summary_writer = tf.summary.create_file_writer(CLS_compare_test_log_dir)
        # self.generator_img_writer = tf.summary.create_file_writer(generator_img_logdir)
        # self.compare_fake_real_img_writer = tf.summary.create_file_writer( "logs/compare_fake_real_img/" +client_name +"/"+self.version_num)

    def __call__(self):
        if self.init is None:
            self.init = tf.group(tf.global_variables_initializer(), tf.local_variables_initializer())
        return self.init

    # @tf.function   # #@tf.function -> cancel eager mode --> able to convert tensor to array
    def train_cls_step(self,batch_data, layer_num_list = None):
        if type(batch_data[0]) != tuple:
            images, labels = batch_data
        else:  #type(batch_data[0]) == tuple  means batch_data contains feature dataset
            images, labels = batch_data[-1]  #last element in batch_data is client train_data
        with tf.GradientTape() as tape:
            loss = 0.0
            predictions = self.cls(images, training=True)
            for teacher in self.teacher_list:
                if self.config.soft_target_loss_weight != float(0):
                    teacher_predictions = teacher(images, training=False)
                    soft_targets = tf.nn.softmax(predictions/self.config.T)
                    soft_prob = tf.nn.log_softmax(teacher_predictions/self.config.T)
                    soft_targets_loss = -tf.math.reduce_sum(soft_targets * soft_prob) / soft_prob.shape[0] * (self.config.T**2)
                    loss += self.config.soft_target_loss_weight * soft_targets_loss

                if self.config.hidden_rep_loss_weight != float(0):
                    teacher_features = teacher.get_features(images)
                    student_feature = self.cls.get_features(images)
                    for layer_num, feature in teacher_features.items():
                        teacher_features = tf.reshape(feature, [-1,])
                        student_feature = tf.reshape(student_feature[layer_num], [-1,])
                        cos_sim = tf.tensordot(teacher_features, student_feature,axes=1)/(tf.linalg.norm(teacher_features)*tf.linalg.norm(student_feature)+0.001)
                        hidden_rep_loss = 1 - cos_sim
                        loss += self.config.hidden_rep_loss_weight * hidden_rep_loss

            predictions = tf.nn.softmax(predictions)
            if self.config.dataset == "elliptic":
                y_true = tf.expand_dims(labels, axis=1)
                predictions = predictions[:,1]
                predictions = tf.expand_dims(predictions, axis=1)
                label_loss = tf.nn.weighted_cross_entropy_with_logits(labels=y_true, logits=predictions, pos_weight=self.weights)
                self.cls_train_elliptic_f1(y_true, predictions)
                self.cls_train_elliptic_precision(y_true, predictions)
                self.cls_train_elliptic_recall(y_true, predictions)
            else: 
                label_loss = self.img_loss_fn_cls(labels, predictions)
            self.cls_train_classify_loss(label_loss)
            loss += label_loss * self.original_cls_loss_weight
            distance_loss = 0.0
            feature_loss, total_feature_loss = 0.0, 0.0
            if self.pre_features_central is not None:
                distance_loss = self.count_features_central_distance(self.pre_features_central, images, labels)
                for layer_num, (features, features_label) in zip(layer_num_list, batch_data[:-1]):
                    feature_predictions = self.cls.call_2(layer_num, features)
                    feature_predictions = tf.nn.softmax(feature_predictions)
                    if self.config.dataset != "elliptic":
                        feature_loss = self.feature_loss_fn_cls(features_label, feature_predictions)
                    else:
                        feature_true = tf.expand_dims(features_label, axis=1)
                        feature_predictions = feature_predictions[:,1]
                        feature_predictions = tf.expand_dims(feature_predictions, axis=1)
                        feature_loss = tf.nn.weighted_cross_entropy_with_logits(labels=feature_true, logits=feature_predictions, pos_weight=self.weights)
                    loss += (feature_loss*self.feat_loss_weight)
                    total_feature_loss += feature_loss
            elif self.initial_feature_center is not None:
                distance_loss = self.count_features_central_distance(self.initial_feature_center, images, labels)
            loss += (distance_loss*self.cos_loss_weight)
        gradients = tape.gradient(loss, self.cls.trainable_variables)
        self.cls_optimizer.apply_gradients(zip(gradients, self.cls.trainable_variables))
        return self.cls_train_loss(loss), self.cls_train_accuracy(labels, predictions), self.cls_train_distance_loss(distance_loss),self.cls_train_feature_loss(total_feature_loss)
    
    @tf.function
    def local_test_cls_step(self,images, labels):
        predictions = self.cls(images, training=False)
        if self.config.dataset == "elliptic":
            y_true = tf.expand_dims(labels, axis=1)
            predictions = tf.nn.softmax(predictions)
            predictions = predictions[:,1]
            predictions = tf.expand_dims(predictions, axis=1)
            loss = tf.nn.weighted_cross_entropy_with_logits(labels=y_true, logits=predictions, pos_weight=self.weights) * self.original_cls_loss_weight
            self.cls_test_elliptic_f1(y_true, predictions)
            self.cls_test_elliptic_precision(y_true, predictions)
            self.cls_test_elliptic_recall(y_true, predictions)
        else:
            loss = self.img_loss_fn_cls(labels, predictions)
        return self.cls_test_loss(loss), self.cls_test_accuracy(labels, predictions)
    
    @tf.function
    def global_test_cls_step(self,images, labels):
        predictions = self.cls(images, training=False)
        if self.config.dataset == "elliptic":
            y_true = tf.expand_dims(labels, axis=1)
            predictions = tf.nn.softmax(predictions)
            predictions = predictions[:,1]
            predictions = tf.expand_dims(predictions, axis=1)
            self.cls_global_test_elliptic_f1(y_true, predictions)
            self.cls_global_test_elliptic_precision(y_true, predictions)
            self.cls_global_test_elliptic_recall(y_true, predictions)
        return self.cls_global_test_accuracy(labels, predictions)
    
    def get_features_central(self, images, labels):  
        #using client train dataset to generate features central
        # feature_avg_dic = {}
        # for label in set(labels):
        #     label_index = np.where(labels==label)
        #     feature = self.cls.get_features(images[label_index])
        #     avg_feature = tf.reduce_mean(feature, axis=0) 
        #     feature_avg_dic[label] = avg_feature
        # return feature_avg_dic
        feature_output_layer_feature_avg_dic = {i:{} for i in self.config.features_ouput_layer_list}
        for label in set(labels):
            label_index = np.where(labels==label)
            feature_list = self.cls.get_features(images[label_index])
            for k,v in feature_list.items():
                avg_feature = tf.reduce_mean(v, axis=0) 
                feature_output_layer_feature_avg_dic[k][label] = avg_feature
        return feature_output_layer_feature_avg_dic

    def count_features_central_distance(self, features_central, images, labels):
        feature = self.cls.get_features(images)
        # labels = np.argmax(labels, axis=1)
        accumulate_loss = 0
        for k,v in feature.items():
            for vector,label in zip(v,labels): 
                pre_vector = features_central[k][label.numpy()]
                vector = tf.reshape(vector, [-1,])
                pre_vector = tf.reshape(pre_vector, [-1,])
                cos_sim = tf.tensordot(vector, pre_vector,axes=1)/(tf.linalg.norm(vector)*tf.linalg.norm(pre_vector)+0.001)
                accumulate_loss += 1 - cos_sim
        return accumulate_loss / len(labels)
        # for vector,label in zip(feature[self.config.features_ouput_layer[0]],labels):  
        #     pre_vector = self.pre_features_central[label]
        #     vector = tf.reshape(vector, [-1,])
        #     pre_vector = tf.reshape(pre_vector, [-1,])
        #     cos_sim = tf.tensordot(vector, pre_vector,axes=1)/(tf.linalg.norm(vector)*tf.linalg.norm(pre_vector)+0.001)
        #     accumulate_loss += 1 - cos_sim
        # return accumulate_loss / len(labels)

    def train_cls(self, worksheet, feature_data, suffix):
        checkpoint_dir = './tmp/%s/%s%s/cls_training_checkpoints/'%(self.client_name, self.version_num,suffix)
        if not os.path.exists(checkpoint_dir):
            os.makedirs(checkpoint_dir+'local')  #store model weight
            os.makedirs(checkpoint_dir+'global')
            os.makedirs(checkpoint_dir+'local_checkpoint')   #store checkpoint
            os.makedirs(checkpoint_dir+'global_checkpoint')
            if self.config.dataset == "elliptic":
                os.makedirs(checkpoint_dir+'global_f1')
        checkpoint_local_prefix = os.path.join(checkpoint_dir+'local_checkpoint', "ckpt")
        checkpoint_global_prefix = os.path.join(checkpoint_dir+'global_checkpoint', "ckpt")
        checkpoint = tf.train.Checkpoint( optimizer=self.cls_optimizer,
                                        cls=self.cls, max_to_keep=tf.Variable(1))
        #read latest_checkpoint
        # checkpoint.restore(tf.train.latest_checkpoint('./tmp/%s/13_with_0/cls_training_checkpoints/'%(self.client_name)+'local')) 
        if feature_data is None:
            if len(self.train_data) >  self.batch_size:
                train_data  = self.train_data.batch(self.batch_size,drop_remainder=True)
            else:
                train_data  = self.train_data.batch(self.batch_size)
        elif not self.config.update_feature_by_epoch:
            all_dataset = list(feature_data.values())
            all_dataset.append(self.train_data)
        for cur_epoch in range(self.cls.cur_epoch_tensor.numpy(), self.config.cls_num_epochs + 1, 1):
            if feature_data:
                if self.config.update_feature_by_epoch:
                    feature_dataset = {}
                    feature_idx = np.random.choice(range(self.config.total_features_num), size=self.train_data_num, replace=True)
                    for k, v in feature_data.items():
                        v = copy.deepcopy(np.array(v, dtype=object)[feature_idx])
                        feature, labels = zip(*v)
                        v = tf.data.Dataset.from_tensor_slices(
                            (np.array(feature), np.array(labels)))#.shuffle(len(labels))
                        feature_dataset[k] = v
                    all_dataset = list(feature_dataset.values())
                    all_dataset.append(self.train_data)
                if len(self.train_data) >  self.batch_size:
                    all_train_data  = tf.data.Dataset.zip(tuple(all_dataset)).shuffle(len(self.train_data)).batch(self.batch_size,drop_remainder=True)
                else:
                    all_train_data  = tf.data.Dataset.zip(tuple(all_dataset)).shuffle(len(self.train_data)).batch(self.batch_size)
                for batch_data in all_train_data:
                    self.train_cls_step(batch_data, feature_data.keys())
            else:
                for batch_data in train_data:
                    self.train_cls_step(batch_data)

            with self.CLS_train_summary_writer.as_default():
                tf.summary.scalar('cls_loss_'+self.client_name, self.cls_train_loss.result(), step=cur_epoch) 
                tf.summary.scalar('cls_accuracy_'+self.client_name, self.cls_train_accuracy.result(), step=cur_epoch)
                tf.summary.scalar('cls_distance_loss_'+self.client_name, self.cls_train_distance_loss.result(), step=cur_epoch) 
                tf.summary.scalar('cls_feature_loss_'+self.client_name, self.cls_train_feature_loss.result(), step=cur_epoch) 
                tf.summary.scalar('cls_classify_loss_'+self.client_name, self.cls_train_classify_loss.result(), step=cur_epoch) 

            self.cls.cur_epoch_tensor.assign_add(1)

            for(X_test, Y_test) in self.test_data:
                self.local_test_cls_step(X_test, Y_test)

            # recoed test result on tensorboard
            with self.CLS_test_summary_writer.as_default():
                tf.summary.scalar('cls_loss_'+self.client_name, self.cls_test_loss.result(), step=cur_epoch)
                tf.summary.scalar('cls_accuracy_'+self.client_name, self.cls_test_accuracy.result(), step=cur_epoch)
                tf.summary.scalar('compare_cls_accuracy_'+self.client_name, self.cls_test_accuracy.result(), step=cur_epoch)
            
            for (X_test, Y_test) in self.all_test_data:
                self.global_test_cls_step(X_test, Y_test)

            with self.CLS_compare_test_acc_summary_writer.as_default():
                tf.summary.scalar('compare_cls_accuracy_'+self.client_name, self.cls_global_test_accuracy.result(), step=cur_epoch)
            
            self.global_cls_acc_list.append(self.cls_global_test_accuracy.result())
            self.local_cls_acc_list.append(self.cls_test_accuracy.result())
            
            if self.cls_global_test_accuracy.result() == max(self.global_cls_acc_list):
                # self.cls.save_weights(checkpoint_global_path.format(epoch=cur_epoch))
                del_list = os.listdir(checkpoint_dir+'global')
                del_list2 = os.listdir(checkpoint_dir+'global_checkpoint')
                for f,f2 in zip(del_list,del_list2):
                    file_path = os.path.join(checkpoint_dir+'global', f)
                    file_path2 = os.path.join(checkpoint_dir+'global_checkpoint', f2)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                    if os.path.isfile(file_path2):
                        os.remove(file_path2)
                checkpoint.save(file_prefix = checkpoint_global_prefix)
                self.cls.save_weights(f"{checkpoint_dir}/global/cp-{cur_epoch:04d}.ckpt")
                #record metric in best global acc into excel
                for col_num,col_value in enumerate([cur_epoch,  self.cls_global_test_accuracy.result(), self.cls_test_accuracy.result()]):
                    worksheet.cell(row=int(self.version_num)+2, column=col_num+11, value = float(col_value))

            if self.cls_test_accuracy.result() == max(self.local_cls_acc_list):
                # self.cls.save_weights(checkpoint_local_path.format(epoch=cur_epoch))
                del_list = os.listdir(checkpoint_dir+'local')
                del_list2 = os.listdir(checkpoint_dir+'local_checkpoint')
                for f,f2 in zip(del_list,del_list2):
                    file_path = os.path.join(checkpoint_dir+'local', f)
                    file_path2 = os.path.join(checkpoint_dir+'local_checkpoint', f2)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                    if os.path.isfile(file_path2):
                        os.remove(file_path2)
                checkpoint.save(file_prefix = checkpoint_local_prefix)
                self.cls.save_weights(f"{checkpoint_dir}/local/cp-{cur_epoch:04d}.ckpt")

                #record metric in best local acc into excel
                for col_num,col_value in enumerate([self.version_num, self.original_cls_loss_weight, self.cos_loss_weight, self.feat_loss_weight, cur_epoch, self.cls_train_accuracy.result(),self.cls_test_accuracy.result(), self.cls_global_test_accuracy.result(),self.cls_train_distance_loss.result()]):
                    worksheet.cell(row=int(self.version_num)+2, column=col_num+1, value = float(col_value))

            if cur_epoch in self.initial_client_ouput_feat_epochs:
                path = f"tmp/{self.client_name}/{self.version_num}{suffix}/assigned_epoch/{cur_epoch}"
                os.makedirs(path)
                self.cls.save_weights(f"{path}/cp-{cur_epoch:04d}.ckpt")
                features_central = self.get_features_central(self.train_x,self.train_y)
                real_features = self.generate_real_features()
                # with open(f"{path}/features_central.pkl","wb") as fp:
                #     pickle.dump(features_central, fp)
                # np.save(f"{path}/real_features",real_features)
                # np.save(f"{path}/features_label",self.train_y)
                for k,v in real_features.items():
                    os.makedirs(f"{path}/{k}_layer_output")
                    with open(f"{path}/{k}_layer_output/features_central.pkl","wb") as fp:
                        pickle.dump(features_central[k], fp)
                    np.save(f"{path}/{k}_layer_output/real_train_features",v)
                    np.save(f"{path}/{k}_layer_output/train_y",self.train_y)

            template = 'Epoch {}, Loss: {}, Accuracy: {}, Test Loss: {}, Test Accuracy: {}, Global Test Accuracy: {}'
            print (template.format(cur_epoch+1,
                                    self.cls_train_loss.result(), 
                                    self.cls_train_accuracy.result()*100,
                                    self.cls_test_loss.result(), 
                                    self.cls_test_accuracy.result()*100,
                                    self.cls_global_test_accuracy.result()*100,))
            if self.config.dataset == "elliptic":
                with self.CLS_train_summary_writer.as_default():
                    tf.summary.scalar('cls_f1_'+self.client_name, self.cls_train_elliptic_f1.result()[0], step=cur_epoch) 
                self.global_cls_f1_list.append(self.cls_global_test_elliptic_f1.result())
                self.global_cls_recall_list.append(self.cls_global_test_elliptic_recall.result())
                self.global_cls_precision_list.append(self.cls_global_test_elliptic_precision.result())
                template = 'Train Recall: {}, Train Precision: {}, Train F1: {}, Global Test Recall: {}, Global Test Precision: {}, Global Test F1: {}'
                print (template.format(self.cls_train_elliptic_recall.result()*100,   #Test Recall: {}, Test Precision: {}, Test F1: {}, 
                                    self.cls_train_elliptic_precision.result()*100,
                                    self.cls_train_elliptic_f1.result()*100,
                                    self.cls_global_test_elliptic_recall.result()*100,
                                    self.cls_global_test_elliptic_precision.result()*100,
                                    self.cls_global_test_elliptic_f1.result()*100,))

            # Reset metrics every epoch
            self.cls_train_loss.reset_states()
            self.cls_test_loss.reset_states()
            self.cls_train_accuracy.reset_states()
            self.cls_test_accuracy.reset_states()
            self.cls_global_test_accuracy.reset_states()
            self.cls_train_distance_loss.reset_states()
            self.cls_train_feature_loss.reset_states()
            self.cls_train_classify_loss.reset_states()
            self.cls_test_elliptic_recall.reset_states()
            self.cls_test_elliptic_f1.reset_states()
            self.cls_test_elliptic_precision.reset_states()
            self.cls_global_test_elliptic_f1.reset_states()
            self.cls_global_test_elliptic_precision.reset_states()
            self.cls_global_test_elliptic_recall.reset_states()
            self.cls_train_elliptic_recall.reset_states()
            self.cls_train_elliptic_f1.reset_states()
            self.cls_train_elliptic_precision.reset_states()

        best_global_acc = max(self.global_cls_acc_list)
        best_local_acc = max(self.local_cls_acc_list)
        max_local_acc_index = self.local_cls_acc_list.index(best_local_acc)
        max_global_acc_index = self.global_cls_acc_list.index(best_global_acc)
        print("max_local_acc_index",max_local_acc_index,"max_local_acc",best_local_acc)
        print("max_global_acc_index",max_global_acc_index,"max_global_acc", best_global_acc)
        method =  "bl"
        if self.config.initial_client:
            if self.config.whether_initial_feature_center:
                method += "_center_init"
        else:
            if self.config.use_initial_model_weight:
                method += "_ours_avg_init"
            if self.cos_loss_weight > 0:
                method += "_cos"
            if self.feat_loss_weight > 0:
                method += "_feat"
                if self.config.update_feature_by_epoch:
                    method += '_self2others_e'
                elif self.config.feature_match_train_data:
                    method += '_others2self'
                else:
                    method += '_self2others'
        if not os.path.exists(f'./tmp/global_metrics_record.xlsx'):
            #create new exccel to record metrics in all clients
            global_workbook = openpyxl.Workbook() 
            global_worksheet = global_workbook.create_sheet("0", 0)
            for col_num,col_index in enumerate(['dataset', 'method', 'epoch','alpha','batch_size','random_seed','data_random_seed','kernel_initializer', 'client number', 'sample_ratio', 'global_acc', 'std'] + list(range(1, self.config.num_clients+1))+ ['global_f1', 'std'] + list(range(1,self.config.num_clients+1))):
                global_worksheet.cell(row=1, column=col_num+1, value = col_index) 
        else: 
            global_workbook = openpyxl.load_workbook(f'./tmp/global_metrics_record.xlsx')
            global_worksheet = global_workbook['0'] 

        for col_num,col_value in enumerate([self.config.dataset, method, cur_epoch, self.config.alpha, self.batch_size,self.config.random_seed, self.config.data_random_seed, self.config.use_same_kernel_initializer, self.config.num_clients, self.config.sample_ratio]):
            global_worksheet.cell(row=int(self.version_num)+2, column=col_num+1, value = col_value)
        client_num = int(self.client_name.split("_")[-1])
        global_worksheet.cell(row=int(self.version_num)+2, column=client_num+12, value = str(round(float(best_global_acc),4)*100)+"%")
        
        if self.config.dataset == "elliptic":
            best_global_f1 = max(self.global_cls_f1_list)
            max_global_f1_index = self.global_cls_f1_list.index(best_global_f1)
            recall = self.global_cls_recall_list[max_global_f1_index]
            precision = self.global_cls_precision_list[max_global_f1_index]
            print(f"best_global_f1_index: {max_global_f1_index}, best_global_f1: {best_global_f1} recall: {recall}, precision: {precision}",)
            global_worksheet.cell(row=int(self.version_num)+2, column=client_num+19, value = str(round(float(best_global_f1),4)*100)+"%")
        global_workbook.save(f'./tmp/global_metrics_record.xlsx')
        
        #load model in best local test acc
        # checkpoint.restore(tf.train.latest_checkpoint(checkpoint_dir+'local')) 
        latest = tf.train.latest_checkpoint(checkpoint_dir+'local')
        self.cls.load_weights(latest)
        return self.get_features_central(self.train_x,self.train_y), self.generate_real_features(), best_global_acc, best_local_acc
    
    #features_version
    def gradient_penalty(self, real_features, generated_features):
        """ features_version
        Calculates the gradient penalty.

        This loss is calculated on an interpolated image
        and added to the discriminator loss.
        """
        # Get the interpolated image
        alpha = tf.random.normal([self.batch_size, 1], 0.0, 1.0)
        diff = generated_features - real_features
        interpolated = real_features + alpha * diff
        with tf.GradientTape() as gp_tape:
            gp_tape.watch(interpolated)
            # 1. Get the discriminator output for this interpolated image.
            if self.GAN_version == "ACGAN":
                pred, pred_class = self.discriminator(interpolated, training=True)
            else:
                pred = self.discriminator(interpolated, training=True)

        # 2. Calculate the gradients w.r.t to this interpolated image.
        grads = gp_tape.gradient(pred, [interpolated])[0]
        # 3. Calculate the norm of the gradients.
        norm = tf.sqrt(tf.reduce_sum(tf.square(grads), axis=[1]))
        gp = tf.reduce_mean((norm - 1.0) ** 2)
        return gp
    def trainGAN(self):
        checkpoint_dir = './gan_training_checkpoints/%s/'%self.client_name
        checkpoint_prefix = os.path.join(checkpoint_dir, "ckpt")
        checkpoint = tf.train.Checkpoint(generator_optimizer=self.disc_optimizer,
                                        discriminator_optimizer=self.gen_optimizer,
                                        generator=self.generator,
                                        discriminator=self.discriminator)
        #read latest_checkpoint
        # checkpoint.restore(tf.train.latest_checkpoint(checkpoint_dir)) 
        for cur_epoch in range(self.generator.cur_epoch_tensor.numpy(), self.config.GAN_num_epochs + 1, 1):
            # train 
            for x,y in self.train_data:
                self.trainGAN_step(x,y)
    
            # recoed training result on tensorboard
            with self.GAN_train_summary_writer.as_default():
                tf.summary.scalar('disc_loss_'+self.client_name, self.disc_train_loss.result(), step=cur_epoch)
                tf.summary.scalar('gen_loss_'+self.client_name, self.gen_train_loss.result(), step=cur_epoch)
                # tf.summary.scalar('disc_binary_accuracy_'+self.client_name, self.disc_train_binary_accuracy.result(), step=cur_epoch)
                # tf.summary.scalar('disc_categorical_accuracy_'+self.client_name, self.disc_train_categorical_accuracy.result(), step=cur_epoch)

            self.generator.cur_epoch_tensor.assign_add(1)
            self.discriminator.cur_epoch_tensor.assign_add(1)
            display.clear_output(wait=True)
            print('cur_epoch',cur_epoch+1)

            # #get generator output img
            # img = self.generate_and_save_images(self.generator,
            #                 cur_epoch + 1,
            #                 self.seed)
            
            # # show img on tensorboard
            # with self.generator_img_writer.as_default():
            #     tf.summary.image("Generate data", img, step=cur_epoch)
            
            with self.compare_fake_real_img_writer.as_default():
                tf.summary.image("compare fake real img_"+self.client_name, self.generate_tsne_images(), step=cur_epoch)
            
            #test
            for(X_test, Y_test) in self.test_data:
                self.testGAN_step(X_test,Y_test)
            # recoed test result on tensorboard
            with self.GAN_test_summary_writer.as_default():
                tf.summary.scalar('disc_loss_'+self.client_name, self.disc_test_loss.result(), step=cur_epoch)
                tf.summary.scalar('gen_loss_'+self.client_name, self.gen_test_loss.result(), step=cur_epoch)
                # tf.summary.scalar('disc_binary_accuracy_'+self.client_name, self.disc_test_binary_accuracy.result(), step=cur_epoch)
                # tf.summary.scalar('disc_categorical_accuracy_'+self.client_name, self.disc_test_categorical_accuracy.result(), step=cur_epoch)

            #print
            template = 'Epoch {}, Generator Loss: {}, Discriminator Loss: {}, Binary Accuracy: {}, Categorical Accuracy: {}, Generator Test Loss: {}, Discriminator Test Loss: {}, Binary Test Accuracy: {}, Categorical Test Accuracy: {},'
            print (template.format(cur_epoch+1,
                                    self.gen_train_loss.result(), 
                                    self.disc_train_loss.result(), 
                                    self.disc_train_binary_accuracy.result()*100,
                                    self.disc_train_categorical_accuracy.result()*100,
                                    self.gen_test_loss.result(), 
                                    self.disc_test_loss.result(),
                                    self.disc_test_binary_accuracy.result()*100,
                                    self.disc_test_categorical_accuracy.result()*100,))
            if cur_epoch != self.config.GAN_num_epochs:
                # Reset metrics every epoch
                self.gen_train_loss.reset_states()
                self.disc_train_loss.reset_states()
                self.disc_train_binary_accuracy.reset_states()
                self.disc_train_categorical_accuracy.reset_states()
                self.gen_test_loss.reset_states()
                self.disc_test_loss.reset_states()
                self.disc_test_binary_accuracy.reset_states()
                self.disc_test_categorical_accuracy.reset_states()

            #Save the model every 50 epochs
            if (cur_epoch + 1) % 50 == 0:
                checkpoint.save(file_prefix = checkpoint_prefix)
        
        # self.trainGAN_step()
        # Generate after the final epoch
        display.clear_output(wait=True)
        ## generate img from generator
        # img = self.generate_and_save_images(self.generator,
        #                         self.config.GAN_num_epochs + 1,
        #                         self.seed)
        ## show generator digital img on tensorboard
        # with self.generator_img_writer.as_default():
        #     tf.summary.image("Generate data", img, step=self.config.GAN_num_epochs + 1)
        
        return self.disc_test_loss.result(), self.gen_test_loss.result(), self.generate_fake_features()
    
    def generate_fake_features(self):
        noise = tf.random.normal([self.config.test_feature_num, self.latent_dim])
        seed = tf.concat(
            [noise, self.train_y[:self.config.test_feature_num]], axis=1
            )
        fake_features = self.generator(seed)
        return fake_features
    
    def generate_real_features(self):
        return self.cls.get_features(self.train_x)
    
    def get_features_label(self):
        return self.train_y
    
    def generate_tsne_images(self):
        noise = tf.random.normal([self.config.test_sample_num, self.latent_dim])
        seed = tf.concat(
            [noise, self.test_y[:self.config.test_sample_num]], axis=1
            )
        fake_features = self.generator(seed)
        real_features = self.cls.get_features(self.test_x[:self.config.test_sample_num])
        ## for imgs
        # compare_input = tf.concat([self.test_x[:self.config.test_sample_num],fake_features],0)
        ## for features
        compare_input = tf.concat([real_features,fake_features],0)
        data = tf.reshape(compare_input,[2*self.config.test_sample_num,-1])

        tsne = TSNE(n_components=2, verbose=1, random_state=self.config.random_seed)
        z = tsne.fit_transform(data)

        df = pd.DataFrame()
        df["comp-1"] = z[:,0]
        df["comp-2"] = z[:,1]
        buf = io.BytesIO()
        #show discriminator output
        df.loc[:self.config.test_sample_num-1,'y'] = "real"
        df.loc[self.config.test_sample_num:,'y'] = "fake"
        #transfor one_hot to interge
        labels = np.argmax(self.test_y[:self.config.test_sample_num], axis=1)
        df.loc[:self.config.test_sample_num-1,"classes"] = labels 
        df.loc[self.config.test_sample_num:,"classes"] = labels
        sns.scatterplot(x="comp-1", y="comp-2", hue=df.classes.tolist(), style=df.y.tolist(),
                        palette=sns.color_palette("hls", 10),
                        data=df)
        plt.savefig(buf, format='png')
        plt.close()
        buf.seek(0)
        # Convert PNG buffer to TF image
        image = tf.image.decode_png(buf.getvalue(), channels=4)
        # Add the batch dimension
        image = tf.expand_dims(image, 0)
        return image
    
    def discriminator_loss(self, real_img, fake_img):
        real_loss = tf.reduce_mean(real_img)
        fake_loss = tf.reduce_mean(fake_img)
        return fake_loss - real_loss

    def generator_loss(self, fake_img):
        return -tf.reduce_mean(fake_img)

    def generate_and_save_images(self,model, epoch, test_input):
        # Notice `training` is set to False.
        # This is so all layers run in inference mode (batchnorm).
        predictions = model(test_input, training=False)
        # Save the plot to a PNG in memory.
        buf = io.BytesIO()

        fig = plt.figure(figsize=(3, 4))

        for i in range(predictions.shape[0]):
            plt.subplot(3, 4, i+1)
            plt.imshow(predictions[i, :, :, 0]  * 255, cmap='gray')#* 127.5 + 127.5, cmap='gray')
            plt.axis('off')

        plt.savefig(buf, format='png')
        plt.close()
        buf.seek(0)
        # Convert PNG buffer to TF image
        image = tf.image.decode_png(buf.getvalue(), channels=4)
        # Add the batch dimension
        image = tf.expand_dims(image, 0)
        return image
  
    ############features version######
    # @tf.function: The below function is completely Tensor Code
    # Good for optimization
    @tf.function
    # Modify Train step for GAN
    def trainGAN_step(self,images, one_hot_labels):
        # images, one_hot_labels = next(self.data.next_batch(self.batch_size))
        #### use cls model to generate real features
        real_features = self.cls.get_features(images)
        if self.GAN_versionV == "CGAN":
            real_features = tf.concat([real_features, one_hot_labels], -1)
        # Train the discriminator first.
        for i in range(self.discriminator_extra_steps):
            noise = tf.random.normal([self.batch_size, self.latent_dim])
            random_vector_labels = tf.concat(
            [noise, one_hot_labels], axis=1
            )
            with tf.GradientTape() as disc_tape:
                generated_features = self.generator(random_vector_labels, training=True)
                # Combine them with real images. Note that we are concatenating the labels
                # with these images here.
                if self.GAN_version == "CGAN":
                    generated_features = tf.concat([generated_features, one_hot_labels], -1)
                    
                combined_features = tf.concat(
                    [real_features, generated_features], axis=0
                )
                # Assemble labels discriminating real from fake images.
                labels = tf.concat(  
                    [tf.ones((self.batch_size, 1)), tf.zeros((self.batch_size, 1))], axis=0
                )
                if self.GAN_version == "ACGAN":
                    double_one_hot_labels = tf.concat(  
                        [one_hot_labels, one_hot_labels], axis=0
                    )
                    predictions, class_labels = self.discriminator(combined_features, training=True)
                    disc_binary_cost = self.loss_fn_gan_binary(labels,predictions)
                    disc_categorical_cost = self.loss_fn_gan_categorical(double_one_hot_labels, class_labels)
        
                    # Calculate the gradient penalty
                    gp = self.gradient_penalty(real_features, generated_features)    
                    # Add the gradient penalty to the original discriminator loss
                    disc_loss = disc_binary_cost +  disc_categorical_cost  +  gp * self.gp_weight
                else: #CGAN
                    predictions = self.discriminator(combined_features, training=True)
                    disc_cost = self.loss_fn_gan_binary(labels, predictions)
                    gp = self.gradient_penalty(real_features, generated_features)     
                    disc_loss = disc_cost + gp * self.gp_weight  

            # Calculate Gradient
            grad_disc = disc_tape.gradient(disc_loss, self.discriminator.trainable_variables)
            self.disc_optimizer.apply_gradients(zip(grad_disc, self.discriminator.trainable_variables))

        # Train the generator
        # Get the latent vector
        noise = tf.random.normal([self.batch_size, self.latent_dim])
        random_vector_labels = tf.concat(
            [noise, one_hot_labels], axis=1
            )
        # Assemble labels that say "all real images". 
        misleading_labels = tf.ones((self.batch_size, 1))
        with tf.GradientTape() as gen_tape, tf.GradientTape() as disc_tape:
            generated_features = self.generator(random_vector_labels, training=True)
            
            if self.GAN_version == "ACGAN":
                fake_predictions, fake_class_labels = self.discriminator(generated_features, training=True)
                gen_binary_loss = self.loss_fn_gan_binary(misleading_labels, fake_predictions) #lead classify to 
                gen_categorical_loss  = self.loss_fn_gan_categorical(one_hot_labels, fake_class_labels) 
                gen_loss = gen_binary_loss + gen_categorical_loss
            else:
                fake_image_and_labels = tf.concat([generated_features, one_hot_labels], -1)
                fake_output = self.discriminator(fake_image_and_labels, training=True)
                gen_loss = self.loss_fn_gan_binary(misleading_labels, fake_output)

        grad_gen = gen_tape.gradient(gen_loss, self.generator.trainable_variables)
        self.gen_optimizer.apply_gradients(zip(grad_gen, self.generator.trainable_variables))

        return self.gen_train_loss.update_state(gen_loss), self.disc_train_loss.update_state(disc_loss), self.disc_train_binary_accuracy(labels, predictions), self.disc_train_categorical_accuracy(double_one_hot_labels, class_labels)

    @tf.function
    def testGAN_step(self,images, one_hot_labels):
        real_features = self.cls.get_features(images)

        noise = tf.random.normal([self.batch_size, self.latent_dim])
        random_vector_labels = tf.concat(
            [noise, one_hot_labels], axis=1
            )

        generated_features = self.generator(random_vector_labels, training=False)
        if self.GAN_version == 'CGAN':
            generated_features = tf.concat([generated_features, one_hot_labels], -1)
            real_features = tf.concat([real_features, one_hot_labels], -1)
        
        combined_features = tf.concat(
            [real_features, generated_features], axis=0
        )

        # Assemble labels discriminating real from fake images.
        labels = tf.concat(  #set real img classify to zero
            [tf.ones((self.batch_size, 1)), tf.zeros((self.batch_size, 1))], axis=0
        )
        misleading_labels = tf.ones((self.batch_size, 1))

        if self.GAN_version == 'ACGAN':
            double_one_hot_labels = tf.concat(  
                        [one_hot_labels, one_hot_labels], axis=0
                    )
            predictions, class_labels  = self.discriminator(combined_features, training=False)
            #using fake img's output to count generator loss
            gen_binary_loss = self.loss_fn_gan_binary(misleading_labels,predictions[self.batch_size:])
            gen_categorical_loss  = self.loss_fn_gan_categorical(one_hot_labels, class_labels[self.batch_size:]) 
            gen_loss = gen_binary_loss + gen_categorical_loss
            
            disc_binary_cost = self.loss_fn_gan_binary(labels, predictions)
            disc_categorical_cost = self.loss_fn_gan_categorical(double_one_hot_labels, class_labels)
            # Calculate the gradient penalty
            gp = self.gradient_penalty(real_features, generated_features)     
            # Add the gradient penalty to the original discriminator loss
            disc_loss = disc_binary_cost + disc_categorical_cost + gp * self.gp_weight  
        else: #cgan
            predictions_combined_images = self.discriminator(combined_features, training=False)
            disc_cost = self.loss_fn_gan_binary(labels, predictions_combined_images)
            gp = self.gradient_penalty(real_features, generated_features)     
            disc_loss = disc_cost + gp * self.gp_weight 
            
            fake_output = self.discriminator(generated_features, training=False)
            gen_loss = self.loss_fn_gan_binary(misleading_labels,fake_output)
            
        return self.gen_test_loss.update_state(gen_loss), self.disc_test_loss.update_state(disc_loss), self.disc_test_binary_accuracy(labels, predictions), self.disc_test_categorical_accuracy(double_one_hot_labels, class_labels)