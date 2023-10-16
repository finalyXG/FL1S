import tensorflow as tf
from models.example_model import Classifier, ClassifierElliptic, C_Discriminator,C_Generator, AC_Discriminator, AC_Generator
from trainers.example_trainer import Trainer
from data_loader.data_generator import DataGenerator
from tensorboard.plugins.hparams import api as hp
import numpy as np
import os
import argparse
import pickle
import random
import openpyxl
import copy

def concatenate_feature_labels(config, tail_path):
    for index, (client_name, client_version) in enumerate(zip(config.features_central_client_name_list, config.features_central_version_list)):
        path = f"./tmp/{client_name}/{client_version}/{tail_path}"
        if index:
            tmp_feature = np.load(f"{path}/real_train_features.npy",allow_pickle=True)
            if config.feature_type == 'fake':
                tmp_feature = np.load(f"/Users/yangingdai/Downloads/GAN_Templtate/tmp/{client_name}/{config.fake_features_version_list[index]}/fake_features.npy",allow_pickle=True)
            tmp_labels = np.load(f"{path}/train_y.npy",allow_pickle=True)
            feature = np.concatenate((feature, tmp_feature),axis=0)
            labels = np.concatenate((labels, tmp_labels),axis=0)
        else:   #when index == 0 initial feature and labels
            feature = np.load(f"{path}/real_train_features.npy",allow_pickle=True)
            if config.feature_type == 'fake':
                feature = np.load(f"/Users/yangingdai/Downloads/GAN_Templtate/tmp/{client_name}/{config.fake_features_version_list[index]}/fake_features.npy",allow_pickle=True)
            labels = np.load(f"{path}/train_y.npy",allow_pickle=True)
    return feature, labels

def model_avg_init(config, cls):
    weight_object = []
    data_amts = len(config.features_central_client_name_list)
    for index, (client_name, client_version) in enumerate(zip(config.features_central_client_name_list, config.features_central_version_list)):
        checkpoint_dir = f'./tmp/{client_name}/{client_version}/cls_training_checkpoints/local/'
        cls.load_weights(tf.train.latest_checkpoint(checkpoint_dir)).expect_partial()
        weight_object.append(copy.deepcopy(cls.weights))
    if data_amts > 1:
        w_avg = copy.deepcopy(weight_object[0])
        for index in range(len(w_avg)):
            w_avg[index] = w_avg[index]/data_amts
            for i in range(1, data_amts):
                w_avg[index] += weight_object[i][index]/data_amts
        cls.set_weights(copy.deepcopy(w_avg))
    return cls

def create_feature_dataset(config, client_data):
    '''
    generate initial client feature to dataset
    '''
    # if config.use_assigned_epoch_feature:
    #     tail_path = f"assigned_epoch/{config.use_assigned_epoch_feature}/"
    # else:
    #     tail_path = f""
    dataset_dict = {}
    indices, feature_idx = None, None
    (train_data, test_data) = client_data
    client_train_data_num = len(train_data)
    for layer_num in config.features_ouput_layer_list:
        if config.use_assigned_epoch_feature:
            tail_path = f"assigned_epoch/{config.use_assigned_epoch_feature}/{layer_num}_layer_output/"
        else:
            tail_path = f"{layer_num}_layer_output/"
        feature, labels = concatenate_feature_labels(config, tail_path)
        config.total_features_num = len(labels)
        feature_dataset = list(zip(feature, labels))
        if config.take_feature_ratio < 1 and indices is None:
            num_feature_keep = int(config.total_features_num * config.take_feature_ratio)
            indices = np.random.permutation(config.total_features_num)[
                    :num_feature_keep
                ]
        if indices:  #take_feature_ratio
            feature_dataset = np.array(feature_dataset, dtype=object)[indices]
        if not config.update_feature_by_epoch and config.feature_match_train_data and feature_idx is None:
            #Convert the number of initial client feature to be the same as client_data
            feature_idx = np.random.choice(range(config.total_features_num), size=client_train_data_num, replace=True)
        if feature_idx is not None:
            feature_dataset = np.array(feature_dataset, dtype=object)[feature_idx]
        if not config.update_feature_by_epoch:
            feature, labels = zip(*feature_dataset)
            feature_dataset = tf.data.Dataset.from_tensor_slices(
                    (np.array(feature), np.array(labels)))#.shuffle(config.total_features_num)
        dataset_dict[layer_num] = feature_dataset
    if not config.update_feature_by_epoch and not config.feature_match_train_data:
        train_data_idx = np.random.choice(range(client_train_data_num), size=config.total_features_num, replace=True)
        train_data = np.array(train_data, dtype=object)[train_data_idx]
    client_data = (train_data, test_data)
    return dataset_dict, client_data

def generate_initial_feature_center(config, y):
    """
    feature_center_dict[feature_layer_num][data_label]
    """
    initial_feature_center = [tf.random.normal([config.feature_dim, 1])]
    feature_center_dict = {}
    for layer_num in config.features_ouput_layer_list:
        for _ in range(config.num_classes-1):
            while True:
                flag = 1
                tmp_feature = tf.random.normal([config.feature_dim, 1])
                for feature in initial_feature_center:
                    feature = tf.reshape(feature, [-1,])
                    tmp_feature = tf.reshape(tmp_feature, [-1,])
                    cos_sim = tf.tensordot(feature, tmp_feature,axes=1)/(tf.linalg.norm(feature)*tf.linalg.norm(tmp_feature)+0.001)
                    if cos_sim > config.initial_feature_center_cosine_threshold:
                        flag = 0
                        break
                if flag == 1:
                    initial_feature_center.append(tmp_feature)
                    break
        feature_center_dict[layer_num] = {label: feature_center for label, feature_center in zip(set(y), initial_feature_center) }
    return feature_center_dict

def clients_main(config):
    data = DataGenerator(config)
    suffix = ""
    pre_features_central = None
    if not config.initial_client:  
        pre_features_central = {i:None for i in config.features_ouput_layer_list}
        for layer_num in config.features_ouput_layer_list:
            for client_name, client_version in zip(config.features_central_client_name_list, config.features_central_version_list):
                suffix += f"_{client_name}_{client_version}" #indicate clients_name version features center
                if config.use_assigned_epoch_feature:
                    path = f'tmp/{client_name}/{client_version}/assigned_epoch/{config.use_assigned_epoch_feature}/{layer_num}_layer_output/'
                else:
                    path = f'tmp/{client_name}/{client_version}/{layer_num}_layer_output/'
                with open(f'{path}/features_central.pkl','rb') as fp: 
                    features_central = pickle.load(fp) #load features_central pre-saved
                if pre_features_central[layer_num]:
                    for k,v in pre_features_central[layer_num].items():
                        pre_features_central[layer_num][k] = tf.stack([pre_features_central[layer_num][k],features_central[k]])
                        pre_features_central[layer_num][k] = tf.reduce_mean(pre_features_central[layer_num][k], axis=0) 
                else:
                    pre_features_central[layer_num] = features_central

    if not os.path.exists(f"tmp/{config.clients_name}"):
        version_num = 0
        #create new exccel to record metrics in cls best local acc
        workbook = openpyxl.Workbook() 
        worksheet = workbook.create_sheet("0", 0)
        for col_num,col_index in enumerate(["version_num","original_cls_loss_weight", 'cos_loss_weight', "feat_loss_weight", "best_local_acc_epoch", 'Train_acc',"Test_local acc", "Test_global acc","Cos_loss"," ", "best_global_acc_epoch", "best_global_acc", "local_acc_in_best_global_acc_epoch"]):
            worksheet.cell(row=1, column=col_num+1, value = col_index) 
    else:
        file_list = next(os.walk(f"./tmp/{config.clients_name}"))[1]   #get all dir in path
        file_list = [int(i.split("_")[0]) for i in file_list] 
        file_list.sort()
        version_num = file_list[-1]+1  #get latest version num + 1
        workbook = openpyxl.load_workbook(f'./tmp/{config.clients_name}/metrics_record.xlsx')
        worksheet = workbook['0'] 

    all_test_x,all_test_y = data.test_x, data.test_y
    client_data = data.clients[config.clients_name]
    HP_BATCH_SIZE = hp.HParam("batch_size", hp.Discrete(config.batch_size_list))
    HP_COS_LOSS_WEIGHT = hp.HParam("cos_loss_weight", hp.Discrete(config.cos_loss_weight_list))
    HP_ORIGINAL_CLS_LOSS_WEIGHT = hp.HParam("original_cls_loss_weight", hp.Discrete(config.original_cls_loss_weight_list))
    HP_FEAT_LOSS_WEIGHT = hp.HParam("feat_loss_weight", hp.Discrete(config.feat_loss_weight_list))
    HP_LEARNING_RATE = hp.HParam("learning_rate", hp.Discrete(config.learning_rate_list)) #hp.RealInterval(0.001, 0.1))
    # HP_GAN_VERSION = hp.HParam("gan_version", hp.Discrete(['ACGAN', 'CGAN']))
    HPARAMS = [
        HP_BATCH_SIZE,
        HP_LEARNING_RATE,
        HP_COS_LOSS_WEIGHT,
        HP_ORIGINAL_CLS_LOSS_WEIGHT,
        HP_FEAT_LOSS_WEIGHT
        # HP_GAN_VERSION,
    ]
    METRICS = [
        hp.Metric(
            "best_global_acc",
            display_name="best_global_acc",
        ),
        hp.Metric(
            "best_local_acc",
            display_name="best_local_acc",
        ),
    ]
    # with tf.summary.create_file_writer(os.path.join(config.logdir,config.clients_name)).as_default():
    #     hp.hparams_config(hparams=HPARAMS, metrics=METRICS)

    # create an instance of the model
    # for model_version in HP_GAN_VERSION.domain.values:
    for batch_size in HP_BATCH_SIZE.domain.values:
        for learning_rate in HP_LEARNING_RATE.domain.values:
            for original_cls_loss_weight in HP_ORIGINAL_CLS_LOSS_WEIGHT.domain.values: 
                for feat_loss_weight in HP_FEAT_LOSS_WEIGHT.domain.values: 
                    for cos_loss_weight in HP_COS_LOSS_WEIGHT.domain.values: 
                        # reset random seed for each client
                        tf.random.set_seed(args.random_seed)
                        np.random.seed(args.random_seed)
                        random.seed(args.random_seed)
                        hparams = {
                            # HP_GAN_VERSION: model_version,
                            HP_BATCH_SIZE: batch_size,
                            HP_LEARNING_RATE: learning_rate,
                            HP_ORIGINAL_CLS_LOSS_WEIGHT: original_cls_loss_weight,
                            HP_COS_LOSS_WEIGHT: cos_loss_weight,
                            HP_FEAT_LOSS_WEIGHT: feat_loss_weight
                        }
                        print({h.name: hparams[h] for h in hparams})
                        hparams = {h.name: hparams[h] for h in hparams}
                        os.makedirs(f"tmp/{config.clients_name}/{version_num}{suffix}")

                        # record hparams and config value in this version
                        record_hparams_file = open(f"./tmp/{config.clients_name}/{version_num}{suffix}/hparams_record.txt", "wt")
                        for key,value in hparams.items():
                            record_hparams_file.write(f"{key}: {value}")
                            record_hparams_file.write("\n")
                        for key,value in vars(config).items():
                            record_hparams_file.write(f"{key}: {value}")
                            record_hparams_file.write("\n")
                        record_hparams_file.close()

                        print('--- Starting trial: %s' % version_num)
                        # with tf.summary.create_file_writer(os.path.join(config.logdir,config.clients_name,str(version_num))).as_default():
                        hp.hparams(hparams)  # record the values used in this trial
                        if config.dataset != "elliptic":
                            cls = Classifier(config)
                            teacher = Classifier(config)
                        else:
                            cls = ClassifierElliptic(config)
                            teacher = ClassifierElliptic(config)
                        feature_data = None
                        initial_feature_center = None
                        teacher_list = []
                        print("feture dimension", config.feature_dim)
                        if feat_loss_weight == float(0):
                            config.feature_match_train_data = 1
                            print("feature_match_train_data:", config.feature_match_train_data)
                        if not config.initial_client:   # get initial_client's features
                            if config.use_initial_model_weight:
                                cls = model_avg_init(config, cls)
                            feature_data, client_data = create_feature_dataset(config, client_data)
                            if config.soft_target_loss_weight != float(0) or config.hidden_rep_loss_weight != float(0):
                                if config.teacher_repeat:
                                    teacher_idx = np.random.choice(range(len(config.features_central_client_name_list)), size=config.teacher_num, replace=True)
                                else:
                                    teacher_idx = np.random.choice(range(len(config.features_central_client_name_list)), size=config.teacher_num, replace=False)
                                print("teacher_idx",teacher_idx)
                                for index in  teacher_idx:
                                    checkpoint_dir = f'./tmp/{config.features_central_client_name_list[index]}/{config.features_central_version_list[index]}/cls_training_checkpoints/local/'
                                    teacher.load_weights(tf.train.latest_checkpoint(checkpoint_dir)).expect_partial()
                                    teacher_list.append(copy.deepcopy(teacher))
                        elif config.whether_initial_feature_center:
                            initial_feature_center = generate_initial_feature_center(config, all_test_y)
                        generator = AC_Generator(config)
                        discriminator = AC_Discriminator(config)

                        trainer = Trainer(config.clients_name, version_num, client_data, all_test_x, all_test_y, initial_feature_center, pre_features_central, cls, teacher_list, discriminator, generator,config,hparams)
                        cls.init_cur_epoch()
                        cur_features_central, real_features, best_global_acc, best_local_acc = trainer.train_cls(worksheet, feature_data, suffix)
                        # features_label = trainer.get_features_label()
                            ### GAN
                            # print("after train cls")
                            # disc_test_loss, gen_test_loss, fake_features = trainer.trainGAN()
                            # tf.summary.scalar("best_global_acc", best_global_acc, step=1)
                            # tf.summary.scalar("best_local_acc", best_local_acc, step=1)

                        # with open(f"tmp/{config.clients_name}/{version_num}{suffix}/features_central.pkl","wb") as fp:
                        #     pickle.dump(cur_features_central, fp)

                        for k,v in cur_features_central.items():
                            os.makedirs(f"tmp/{config.clients_name}/{version_num}{suffix}/{k}_layer_output")
                            with open(f"tmp/{config.clients_name}/{version_num}{suffix}/{k}_layer_output/features_central.pkl","wb") as fp:
                                pickle.dump(v, fp)
                        # np.save(f"tmp/{config.clients_name}/{version_num}{suffix}/real_features",real_features)
                        # np.save(f"tmp/{config.clients_name}/{version_num}{suffix}/features_label",features_label)
                        version_num += 1
                
    workbook.save(f'./tmp/{config.clients_name}/metrics_record.xlsx')

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    # General command line arguments for all models
    parser.add_argument(
        "--dataset",
        type=str,
        default="mnist"
    )
    parser.add_argument(
        "--clients_name",
        type=str,
        help="Name of client",
        default="clients_1",
    )
    parser.add_argument(
        "--sample_ratio",
        type=float,
        help="Fraction of training data to make available to users",
        default=1,
    )
    parser.add_argument(
        "--take_feature_ratio",
        type=float,
        help="Fraction of pre-feature data to make available to users",
        default=1,
    )
    parser.add_argument(
        "--alpha",
        type=float,
        help="Measure of heterogeneity (higher is more homogeneous, lower is more heterogenous)",
        default=10,
    )
    parser.add_argument("--initial_client", type=int, default=1)  #use 0 and 1 to replace False and True 
    parser.add_argument(
        "--initial_client_ouput_feat_epochs", 
        type=int, 
        nargs='+', 
        help="Define generate trained feature data in which epoch",
        default=[-1]
        ) 
    parser.add_argument(
        "--whether_initial_feature_center", 
        type=int, 
        help="Whether send random initial feature center to initial client",
        default=0
        ) 
    parser.add_argument(
        "--initial_feature_center_cosine_threshold",
        type=float, 
        help="Define the consine similarity socre each initial feature center have to reach",
        default=0.5
    )
    parser.add_argument("--T", type=float, default=1.0)
    parser.add_argument("--soft_target_loss_weight", type=float, default=0.0) 
    parser.add_argument("--hidden_rep_loss_weight", type=float, default=0.0)
    parser.add_argument("--teacher_num", type=int, default=1)
    parser.add_argument("--teacher_repeat", type=int, default=0)
    parser.add_argument("--features_central_client_name_list", type=str, nargs='+', default=["clients_1"])  #use which version as initial client( only for initial_client==0)
    parser.add_argument("--features_central_version_list", type=str,nargs='+',  default=["0"])  #use which version as initial client( only for initial_client==0)
    parser.add_argument(
        "--feature_type",
        type=str,
        help="choose to usereal or fake feature",
        default="real",
    )
    parser.add_argument("--fake_features_version_list", type=str,nargs='+',  default=["0"])  #use which version as initial client( only for initial_client==0)
    parser.add_argument("--use_initial_model_weight", type=int, default=0)  #use which version as initial client( only for initial_client==0)
    parser.add_argument("--use_assigned_epoch_feature", type=int, default=0)  #use 0 means False==> use feature in best local acc( only for initial_client==0)
    parser.add_argument("--use_dirichlet_split_data", type=int, default=1)  #use 0 means False, 1 means True
    parser.add_argument("--use_same_kernel_initializer", type=int, default=1)
    parser.add_argument("--feature_match_train_data", type=int, default=1)  #1 means set the length of feature to be the same as the length of train data, 0 reverse
    parser.add_argument("--update_feature_by_epoch", type=int, default=0)
    parser.add_argument("--cls_num_epochs", type=int, default=20)

    parser.add_argument("--original_cls_loss_weight_list", type=float, nargs='+', default=[1.0])
    parser.add_argument("--feat_loss_weight_list", type=float, nargs='+', default=[1.0])
    parser.add_argument("--cos_loss_weight_list", type=float, nargs='+', default=[5.0]) 
    parser.add_argument("--learning_rate_list", type=float, nargs='+', default=[0.001]) 
    parser.add_argument("--batch_size_list", type=int, nargs='+', default=[32]) 

    parser.add_argument("--features_ouput_layer_list", help="The index of features output Dense layer",nargs='+',type=int, default=[-2])
    parser.add_argument("--GAN_num_epochs", type=int, default=1)
    parser.add_argument("--test_feature_num", type=int, default=500)
    parser.add_argument("--test_sample_num", help="The number of real features and fake features in tsne img", type=int, default=500) 
    parser.add_argument("--gp_weight", type=int, default=10.0)
    parser.add_argument("--discriminator_extra_steps", type=int, default=3)
    parser.add_argument("--num_examples_to_generate", type=int, default=16)
    parser.add_argument("--latent_dim", type=int, default=16)
    parser.add_argument("--feature_dim", type=int, default=128)
    parser.add_argument("--max_to_keep", type=int, default=5)

    parser.add_argument("--num_clients", type=int, default=2)
    parser.add_argument("--client_train_num", type=int, default=1000)
    parser.add_argument("--client_test_num", type=int, default=1000)
    parser.add_argument("--random_seed", type=int, default=10)
    parser.add_argument("--data_random_seed", type=int, default=1693)

    parser.add_argument("--exp_name", type=str, default="example")
    parser.add_argument("--logdir", type=str, default="logs/hparam_tuning")

    args = parser.parse_args()
    args.initial_client = bool(args.initial_client)
    args.use_initial_model_weight = bool(args.use_initial_model_weight)
    args.use_dirichlet_split_data = bool(args.use_dirichlet_split_data)
    args.update_feature_by_epoch = bool(args.update_feature_by_epoch)
    args.teacher_repeat = bool(args.teacher_repeat)
    print("client:", args.clients_name)
    print("Whether initial_client:", args.initial_client)
    print("features_central_version_list:", args.features_central_version_list)
    if not args.use_dirichlet_split_data:
        print("client_train_num:", args.client_train_num)
        print("client_test_num:", args.client_test_num)
    print("cls_num_epochs:", args.cls_num_epochs)
    print("initial_client_ouput_feat_epoch:", args.initial_client_ouput_feat_epochs)
    print("features_ouput_layer_list:",args.features_ouput_layer_list)
    print("use_dirichlet_split_data",args.use_dirichlet_split_data)
    if args.initial_client_ouput_feat_epochs[0] <= args.cls_num_epochs:
        clients_main(args)
    else:
        print("initial_client_ouput_feat_epochs must be smaller than cls_num_epochs")