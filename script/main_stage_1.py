
import tensorflow as tf
from tensorboard.plugins.hparams import api as hp
import numpy as np
import os
import argparse
import pickle
import random
import openpyxl
import copy
import time, datetime
from data_loader.data_generator import DataGenerator
from script.model import Classifier, GAN, reduction_number, reduction_rate, epochs, score0_target1_num, smaller_half_number
class CustomCallback(tf.keras.callbacks.Callback):
    def on_epoch_end(self, epoch, logs=None):
        #save feature, feature center in cur_epoch model 
        if epoch in self.model.config.initial_client_ouput_feat_epochs:
            path = f"script_tmp/stage_1/{model.config.config.dataset}/{model.config.config.clients_name}/{model.config.version_num}/assigned_epoch/{epoch}"
            if not os.path.exists(path):
                os.makedirs(path)
            model.save_weights(f"{path}/cp-{epoch:04d}.ckpt")
            if self.model.config.dataset != 'real_data':
                real_features = self.model.get_features(model.train_x)
                for feature_type, feature in real_features:
                    np.save(f"{path}/real_features_{feature_type}",feature)   #save feature as a dict
                np.save(f"{path}/label",model.train_y)
            else:
                feature, label = model.get_features(train_data)
                if "before_attention" in feature.keys():
                    np.save(f"script_tmp/stage_1/{model.config.dataset}/{model.config.clients_name}/{model.config.version_num}/real_features_BFA",feature['before_attention'])
                if "after_attention" in feature.keys():
                    np.save(f"script_tmp/stage_1/{model.config.dataset}/{model.config.clients_name}/{model.config.version_num}/real_features_AFA",feature['after_attention'])
                np.save(f"{path}/label",label)

        model.epochs.assign(epoch)
        for metric in model.metrics:
            metric.reset_states()
        for metric in model.compiled_metrics._metrics:
            metric.reset_states()

    def on_test_begin(self, logs=None):
        for metric in model.metrics:
            metric.reset_states()
        for metric in model.compiled_metrics._metrics:
            metric.reset_states()
class GANCallback(tf.keras.callbacks.Callback):
    def on_epoch_end(self, epoch, logs=None):
        for metric in model.metrics:
            metric.reset_states()
        for metric in model.compiled_metrics._metrics:
            metric.reset_states()

    def on_test_begin(self, logs=None):
        for metric in model.metrics:
            metric.reset_states()
        for metric in model.compiled_metrics._metrics:
            metric.reset_states()

class LossAndErrorPrintingCallback(tf.keras.callbacks.Callback):
    def on_epoch_end(self, epoch, logs=None):
        print()
        print(f"Epoch: {epoch} ", end='')
        for k,v in logs.items():
            if 'val' not in k:
                if "f1score" in k:
                    print(f"{k} is {v[0]:.5f}, ", end='')
                else:
                    print(f"{k} is {v:.5f}, ", end='')
        print()

    def on_test_end(self, logs=None):
        print("test metrics:", end='')
        for k,v in logs.items():
            if "f1score" in k:
                print(f"{k} is {v[0]:.5f}, ", end='')
            else:
                print(f"{k} is {v:.5f}, ", end='')
        print()

class RecordTableCallback(tf.keras.callbacks.Callback):
    def on_test_end(self, logs=None):
        if not os.path.exists(f'script_tmp/stage_1/metrics_record.xlsx'):
            #create new exccel to record metrics in all clients
            global_workbook = openpyxl.Workbook() 
            global_worksheet = global_workbook.create_sheet("0", 0)
            for col_num,col_index in enumerate(["current_time",'dataset',"clients_name","temperature","loss_weight","loss", 'acc', "rd", "rr", 'f1', 'recall', 'precision']):
                global_worksheet.cell(row=1, column=col_num+1, value = col_index) 
        else: 
            global_workbook = openpyxl.load_workbook(f'script_tmp/stage_1/metrics_record.xlsx')
            global_worksheet = global_workbook['0']
        global_worksheet.append([datetime.datetime.now().strftime("%Y%m%d-%H%M%S"), model.config.dataset,model.config.clients_name, model.config.temperature, model.config.client_loss_weight, float(logs['loss']),float(logs['cls_accuracy']),int(logs['reduction_number']),float(logs['reduction_rate']),float(logs['f1score']),float(logs['recall']),float(logs['precision'])])
        global_workbook.save('script_tmp/stage_1/metrics_record.xlsx')

def generate_initial_feature_center(config):
    """
    feature_center_dict[feature_layer_num][data_label]
    """
    initial_feature_center = [tf.random.normal([config.feature_dim, 1])]
    feature_center_dict = {}
    for layer_num in config.features_ouput_layer_list:
        for _ in range(config.num_classes-1):
            while True:
                flag = 1
                tmp_feature = tf.random.normal([config.feature_dim, 1])
                for feature in initial_feature_center:
                    feature = tf.reshape(feature, [-1,])
                    tmp_feature = tf.reshape(tmp_feature, [-1,])
                    cos_sim = tf.tensordot(feature, tmp_feature,axes=1)/(tf.linalg.norm(feature)*tf.linalg.norm(tmp_feature)+0.001)
                    if cos_sim > config.initial_feature_center_cosine_threshold:
                        flag = 0
                        break
                if flag == 1:
                    initial_feature_center.append(tmp_feature)
                    break
        feature_center_dict[layer_num] = {label: feature_center for label, feature_center in zip(range(config.num_classes), initial_feature_center)}
    return feature_center_dict

def change_dict_architecture(x):
    for q in x['tx'].keys():
        x[f'tx.{q}'] = x['tx'][q]
    del x['tx']
    for q in x['pp'].keys():
        x[f'pp.{q}'] = x['pp'][q]
    del x['pp']
    return x

def generate_dataset(x):
    def gen():
        for i in x:
            yield {'segment':tf.constant(i['segment']),
                'family':tf.constant(i['family']),
                'target':tf.constant(i['target']),
                'tx':{
                    'country':tf.ragged.constant([i['tx']['country']]),
                    'transaction': tf.ragged.constant([i['tx']['transaction']]),
                    'amount': tf.ragged.constant([i['tx']['amount']])
                    },
                'pp':{
                    'p1':tf.ragged.constant([i['pp']['p1']]),
                    'p2': tf.ragged.constant([i['pp']['p2']]),
                    }
                }
    dataset = tf.data.Dataset.from_generator(
        gen,
        output_signature=(
            {
                'segment': tf.TensorSpec(shape=(), dtype=tf.string),
                'family': tf.TensorSpec(shape=(), dtype=tf.string),
                'target': tf.TensorSpec(shape=(), dtype=tf.int32),
                'tx': {
                    'country': tf.RaggedTensorSpec(shape=(1, None), dtype=tf.string),
                    'transaction': tf.RaggedTensorSpec(shape=(1, None), dtype=tf.string),
                    'amount': tf.RaggedTensorSpec(shape=(1, None), dtype=tf.float64),
                },
                'pp': {
                    'p1': tf.RaggedTensorSpec(shape=(1, None), dtype=tf.float64),
                    'p2': tf.RaggedTensorSpec(shape=(1, None), dtype=tf.string),
                }
            }
        )
    )
    return dataset
        
def main(config, model, train_data, test_data, global_test_data):
    tf.random.set_seed(config.random_seed)
    np.random.seed(config.random_seed)
    random.seed(config.random_seed)
    if config.dataset != 'real_data':
        train_x, train_y = zip(*train_data)
        model.set_train_x_train_y(train_x, train_y)
        class_rate = train_y.count(0)/train_y.count(1)
        test_x, test_y = zip(*test_data)
        global_test_x, global_test_y  = zip(*global_test_data)
        train_x, train_y = np.array(train_x),np.array(train_y)
        test_x, test_y = np.array(test_x),np.array(test_y)
        global_test_x, global_test_y = np.array(global_test_x),np.array(global_test_y)

    if not os.path.exists(f"script_tmp/stage_1/{args.dataset}/{config.clients_name}"):
        version_num = 0
    else:
        file_list = next(os.walk(f"./script_tmp/stage_1/{args.dataset}/{config.clients_name}"))[1]   #get all dir in path
        file_list = [int(i.split("_")[0]) for i in file_list] 
        file_list.sort()
        version_num = file_list[-1]+1  #get latest version num + 1
    model.set_version_num(version_num)

    os.makedirs(f"./script_tmp/stage_1/{args.dataset}/{args.clients_name}/{version_num}/")
    record_hparams_file = open(f"./script_tmp/stage_1/{args.dataset}/{args.clients_name}/{version_num}/hparams_record.txt", "wt")
    for key,value in vars(args).items():
        record_hparams_file.write(f"{key}: {value}")
        record_hparams_file.write("\n")
    record_hparams_file.close()

    if config.dataset == "elliptic" or config.dataset == "real_data":
        metrics_list = [reduction_number(), reduction_rate(), epochs(), score0_target1_num(), smaller_half_number(),
                        tf.keras.metrics.BinaryAccuracy(name='cls_accuracy'),
                        tf.keras.metrics.Recall(name='recall'),
                        tf.keras.metrics.Precision(name='precision'),
                        tf.keras.metrics.F1Score(threshold=0.5, name='f1score')]
        if config.dataset == "elliptic":
             model.set_loss_weight(class_rate)
    else:
        metrics_list = [tf.keras.metrics.SparseCategoricalAccuracy(name='cls_accuracy')]
        
    if config.model_save_metrics == "acc":
        monitor = 'val_cls_accuracy'
    elif config.model_save_metrics == "f1":
        monitor = 'val_f1score'
    elif config.model_save_metrics == "rd":
        monitor = 'val_reduction_number'
    elif config.model_save_metrics == "rr":
        monitor = 'val_reduction_rate'
    model.compile(optimizer=tf.keras.optimizers.legacy.Adam(config.learning_rate),
                    metrics= metrics_list,
                    loss=tf.keras.metrics.Mean(name='loss'),
                    run_eagerly = True)

    checkpoint_filepath = f'./script_tmp/stage_1/{config.dataset}/{config.clients_name}/{version_num}/checkpoint/checkpoint'
    model_checkpoint_callback = tf.keras.callbacks.ModelCheckpoint(
        filepath=checkpoint_filepath,
        save_weights_only=True,
        monitor=monitor,
        mode='max',
        save_best_only=True)
    early_stopping = tf.keras.callbacks.EarlyStopping(monitor=monitor, patience=8, mode="max")
    if config.validation_data == "local_test_data":
        if config.dataset == "real_data":
            validation_data = test_data
            validation_batch_size = config.client_test_size
        else:
            validation_data = (test_x, test_y)
            validation_batch_size = len(test_y)
    elif config.validation_data == "global_test_data":
        if config.dataset == "real_data":
            validation_data = global_test_data
            validation_batch_size = config.total_test_size
        else:
            validation_data = (global_test_x, global_test_y)
            validation_batch_size = len(global_test_y)

    log_dir = "script_logs/fit/" + datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    tensorboard_callback = tf.keras.callbacks.TensorBoard(log_dir=log_dir)
    if config.dataset != "real_data":
        if len(train_y) >  config.batch_size:
            train_data_batched = tf.data.Dataset.from_tensor_slices(
                (train_x,train_y)).shuffle(len(train_y)).batch(config.batch_size,drop_remainder=True)
        else:
            train_data_batched = tf.data.Dataset.from_tensor_slices(
                (train_x,train_y)).shuffle(len(train_y)).batch(config.batch_size)
        history = model.fit(train_data_batched, epochs=config.cls_num_epochs, verbose=0, shuffle=True, validation_data=validation_data, validation_batch_size = validation_batch_size, callbacks=[early_stopping, tensorboard_callback, CustomCallback(), model_checkpoint_callback, LossAndErrorPrintingCallback() ])
        test_score = model.evaluate(validation_data, batch_size = validation_batch_size, callbacks=[LossAndErrorPrintingCallback(),CustomCallback()],return_dict=True, verbose=0)

        model.load_weights(checkpoint_filepath)
        test_score = model.evaluate(validation_data, batch_size = validation_batch_size, callbacks=[RecordTableCallback(), LossAndErrorPrintingCallback(),CustomCallback()],return_dict=True, verbose=0)
        feature_dict = model.get_features(train_x)
        for feature_type, feature in feature_dict:
            np.save(f"script_tmp/stage_1/{config.dataset}/{config.clients_name}/{version_num}/real_features_{feature_type}",feature)   #save feature as a dict
        np.save(f"script_tmp/stage_1/{config.dataset}/{config.clients_name}/{version_num}/label",train_y)
    else:
        train_data_batched = train_data.batch(config.batch_size)
        validation_data_batched = validation_data.batch(validation_batch_size)

        history = model.fit(train_data_batched, epochs=config.cls_num_epochs, verbose=0, shuffle=True, validation_data=validation_data_batched, callbacks=[early_stopping, tensorboard_callback, CustomCallback(), model_checkpoint_callback, LossAndErrorPrintingCallback() ])
        test_score = model.evaluate(validation_data_batched, callbacks=[LossAndErrorPrintingCallback(),CustomCallback()],return_dict=True, verbose=0)

        model.load_weights(checkpoint_filepath)
        test_score = model.evaluate(validation_data_batched, callbacks=[RecordTableCallback(), LossAndErrorPrintingCallback(),CustomCallback()],return_dict=True, verbose=0)

        feature, label = model.get_features(train_data)
        if "before_attention" in feature.keys():
            np.save(f"script_tmp/stage_1/{config.dataset}/{config.clients_name}/{version_num}/real_features_BFA",feature['before_attention'])
        if "after_attention" in feature.keys():
            np.save(f"script_tmp/stage_1/{config.dataset}/{config.clients_name}/{version_num}/real_features_AFA",feature['after_attention'])
        
        np.save(f"script_tmp/stage_1/{config.dataset}/{config.clients_name}/{version_num}/label",label)
        config.feature_space.save(f"script_tmp/stage_1/{config.dataset}/{config.clients_name}/{version_num}/featurespace.keras")
    
    #using GAN
    if config.dataset == "real_data" :
        feature = feature[config.fake_data_feature_type]
        config.feature_dim = feature.shape[1]
        if len(label) >  config.batch_size:
            train_data_batched = tf.data.Dataset.from_tensor_slices(
                (feature, label )).shuffle(len(label)).batch(config.batch_size,drop_remainder=True)
        else:
            train_data_batched = tf.data.Dataset.from_tensor_slices(
                (feature, label )).shuffle(len(label)).batch(config.batch_size)
            
        validation_feature, validation_label = model.get_features(validation_data)
        validation_feature = validation_feature[config.fake_data_feature_type]

        if len(validation_label) >  config.batch_size:
            validation_data = tf.data.Dataset.from_tensor_slices(
                (validation_feature, validation_label)).shuffle(len(validation_label)).batch(config.batch_size,drop_remainder=True)
        else:
            validation_data = tf.data.Dataset.from_tensor_slices(
                (validation_feature, validation_label)).shuffle(len(validation_label)).batch(config.batch_size)
        
    if config.whether_generate_fake_feature:
        if config.dataset == "elliptic" or config.dataset == "real_data" :
            label_dim = 1  # for BinaryCrossentropy
            loss_fn_categorical = tf.keras.losses.BinaryCrossentropy(from_logits=True)
        else:
            label_dim = config.num_classes
            loss_fn_categorical = tf.keras.losses.CategoricalCrossentropy(from_logits=True)
        discriminator = tf.keras.Sequential(
            [
                tf.keras.layers.Input(shape=(config.feature_dim)),
                tf.keras.layers.Flatten(), 
                tf.keras.layers.Dense(512, activation=tf.nn.leaky_relu ),
                tf.keras.layers.Dense(256, activation=tf.nn.leaky_relu ),
                tf.keras.layers.Dense(128, activation=tf.nn.leaky_relu ),
                tf.keras.layers.Dense(1+label_dim),
            ],
            name="discriminator",
        )

        # Create the generator
        generator = tf.keras.Sequential(
            [
                tf.keras.layers.Input(shape=(config.latent_dim+1,)),
                tf.keras.layers.Dense(32, activation=tf.nn.relu),
                tf.keras.layers.Dense(64, activation=tf.nn.relu),
                tf.keras.layers.Dense(128, activation=tf.nn.relu),
                tf.keras.layers.Dense(config.feature_dim),
            ],
            name="generator",
        )
        gan = GAN(config=config, discriminator=discriminator, generator=generator, model = model)

        gan.compile(
            d_optimizer=tf.keras.optimizers.legacy.Adam(learning_rate=config.learning_rate),
            g_optimizer=tf.keras.optimizers.legacy.Adam(learning_rate=config.learning_rate),
            loss_fn_binary= tf.keras.losses.BinaryCrossentropy(from_logits=True),
            loss_fn_categorical = loss_fn_categorical,
            binary_accuracy = tf.keras.metrics.BinaryAccuracy(name='disc_binary_accuracy'), 
            categorical_accuracy = tf.keras.metrics.CategoricalAccuracy(name='disc_categorical_accuracy'),
            run_eagerly = True
        )
        gan_checkpoint_filepath = f'script_tmp/stage_1/{config.dataset}/{config.clients_name}/{version_num}/gan_checkpoint/checkpoint'
        gan_checkpoint_callback = tf.keras.callbacks.ModelCheckpoint(
            filepath=gan_checkpoint_filepath,
            save_weights_only=True,
            monitor='val_g_loss',
            mode='min',
            save_best_only=True)
        gan.fit(train_data_batched, epochs=config.GAN_num_epochs, verbose=1, shuffle=True, validation_data=validation_data, callbacks=[GANCallback(), gan_checkpoint_callback, LossAndErrorPrintingCallback()])
        gan.evaluate(validation_data, callbacks=[LossAndErrorPrintingCallback(),GANCallback()],return_dict=True, verbose=0)

        gan.load_weights(gan_checkpoint_filepath)
        gan.evaluate(validation_data, callbacks=[LossAndErrorPrintingCallback(),GANCallback()],return_dict=True, verbose=0)
        fake_features = gan.generate_fake_features(label)
        if "before_attention"  == config.fake_data_feature_type:
            np.save(f"script_tmp/stage_1/{config.dataset}/{config.clients_name}/{version_num}/fake_features_BFA",fake_features)
        elif "after_attention" == config.fake_data_feature_type:
            np.save(f"script_tmp/stage_1/{config.dataset}/{config.clients_name}/{version_num}/fake_features_AFA",fake_features)
        else:
            np.save(f"script_tmp/stage_1/{config.dataset}/{config.clients_name}/{version_num}/fake_features",fake_features)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    # General command line arguments for all models
    parser.add_argument(
        "--stage",
        type=int,
        default=1
    )
    parser.add_argument(
        "--dataset",
        type=str,
        default="mnist"
    )
    parser.add_argument(
        '--model_type',
        type=str,
        default='MHA01',
        help='''
        MHA01: use MHA for each subsequence;
        MHA02: use MHA for each type of sequence
        '''
    )
    parser.add_argument(
        "--disable_lv1_namaes",
        type=str,
        nargs='+',
        default=[""]
    )
    parser.add_argument(
        "--disable_pp_namaes",
        type=str,
        nargs='+',
        default=[""]
    )
    parser.add_argument(
        "--disable_tx_namaes",
        type=str,
        nargs='+',
        default=[""]
    )
    parser.add_argument(
        "--real_data_feature_type",
        type=str,
        nargs='+',
        default=['before_attention'],
        help='''
        before_attention: BFA
        after_attention: AFA
        '''
    )
    parser.add_argument(
        "--fake_data_feature_type",
        type=str,
        default='before_attention',
        help='''
        before_attention
        after_attention
        '''
    )
    parser.add_argument(
        "--input_feature_overlap_num",
        type=int,
        default=165
    )
    parser.add_argument(
        "--temperature",
        type=float,
        default=1.0
    )
    parser.add_argument(
        "--dropout_rate",
        type=float,
        default=0.2
    )
    parser.add_argument(
        "--skip_if_all_0",
        type=int,
        default=0
    )
    parser.add_argument(
        "--whether_use_transformer_model",
        type=int,
        default=0
    )
    parser.add_argument(
        "--client_loss_weight",
        type=float,
        default=0
    )
    parser.add_argument(
        "--gamma",
        type=float,
        default=2.0
    )
    parser.add_argument(
        "--input_feature_no_overlap_num",
        type=int,
        default=0
    )
    parser.add_argument(
        "--extend_null_feature_num",
        type=int,
        default=0
    )
    parser.add_argument(
        "--validation_data",
        type=str,
        default="local_test_data"
    )
    parser.add_argument(
        "--whether_generate_fake_feature",
        type=int,
        default=0
    )
    parser.add_argument(
        "--clients_name",
        type=str,
        help="Name of client",
        default="clients_1",
    )
    parser.add_argument(
        "--sample_ratio",
        type=float,
        help="Fraction of training data to make available to users",
        default=1,
    )
    parser.add_argument(
        "--take_feature_ratio",
        type=float,
        help="Fraction of pre-feature data to make available to users",
        default=1,
    )
    parser.add_argument(
        "--alpha",
        type=float,
        help="Measure of heterogeneity (higher is more homogeneous, lower is more heterogenous)",
        default=10,
    )
    parser.add_argument("--initial_client", type=int, default=1)  #use 0 and 1 to replace False and True 
    parser.add_argument(
        "--initial_client_ouput_feat_epochs", 
        type=int, 
        nargs='+', 
        help="Define generate trained feature data in which epoch",
        default=[-1]
        ) 
    parser.add_argument(
        "--whether_initial_feature_center", 
        type=int, 
        help="Whether send random initial feature center to initial client",
        default=0
        ) 
    parser.add_argument(
        "--initial_feature_center_cosine_threshold",
        type=float, 
        help="Define the consine similarity socre each initial feature center have to reach",
        default=0.5
    )
    parser.add_argument( "--model_save_metrics", type=str, default="acc")
    parser.add_argument("--T", type=float, default=1.0)
    parser.add_argument("--soft_target_loss_weight", type=float, default=0.0) 
    parser.add_argument("--hidden_rep_loss_weight", type=float, default=0.0)
    parser.add_argument("--teacher_num", type=int, default=1)
    parser.add_argument("--teacher_repeat", type=int, default=0)
    parser.add_argument("--features_central_client_name_list", type=str, nargs='+', default=["clients_1"])  #use which version as initial client( only for initial_client==0)
    parser.add_argument("--features_central_version_list", type=str,nargs='+',  default=["0"])  #use which version as initial client( only for initial_client==0)
    parser.add_argument(
        "--feature_type",
        type=str,
        help="choose to usereal or fake feature",
        default="real",
    )
    parser.add_argument("--fake_features_version_list", type=str,nargs='+',  default=["0"])  #use which version as initial client( only for initial_client==0)
    parser.add_argument("--use_initial_model_weight", type=int, default=0)  #use which version as initial client( only for initial_client==0)
    parser.add_argument("--use_assigned_epoch_feature", type=int, default=0)  #use 0 means False==> use feature in best local acc( only for initial_client==0)
    parser.add_argument("--use_dirichlet_split_data", type=int, default=1)  #use 0 means False, 1 means True
    parser.add_argument("--use_same_kernel_initializer", type=int, default=1)
    parser.add_argument("--feature_match_train_data", type=int, default=1)  #1 means set the length of feature to be the same as the length of train data, 0 reverse
    parser.add_argument("--update_feature_by_epoch", type=int, default=0)
    parser.add_argument("--cls_num_epochs", type=int, default=20)

    parser.add_argument("--original_cls_loss_weight_list", type=float, default=1.0)
    parser.add_argument("--feat_loss_weight_list", type=float, default=1.0)
    parser.add_argument("--cos_loss_weight", type=float, default=5.0) 
    parser.add_argument("--learning_rate", type=float,default=0.001) 
    parser.add_argument("--batch_size", type=int, default=32) 

    parser.add_argument("--features_ouput_layer_list", help="The index of features output Dense layer",nargs='+',type=int, default=[-2])
    parser.add_argument("--GAN_num_epochs", type=int, default=1)
    parser.add_argument("--test_feature_num", type=int, default=500)
    parser.add_argument("--test_sample_num", help="The number of real features and fake features in tsne img", type=int, default=500) 
    parser.add_argument("--gp_weight", type=int, default=10.0)
    parser.add_argument("--discriminator_extra_steps", type=int, default=3)
    parser.add_argument("--num_examples_to_generate", type=int, default=16)
    parser.add_argument("--latent_dim", type=int, default=16)
    parser.add_argument("--feature_dim", type=int, default=128)
    parser.add_argument("--max_to_keep", type=int, default=5)

    parser.add_argument("--num_clients", type=int, default=2)
    parser.add_argument("--client_train_num", type=int, default=1000)
    parser.add_argument("--client_test_num", type=int, default=1000)
    parser.add_argument("--random_seed", type=int, default=10)
    parser.add_argument("--data_random_seed", type=int, default=1693)

    parser.add_argument("--exp_name", type=str, default="example")
    parser.add_argument("--logdir", type=str, default="logs/hparam_tuning")

    args = parser.parse_args()
    args.initial_client = bool(args.initial_client)
    args.use_initial_model_weight = bool(args.use_initial_model_weight)
    args.use_dirichlet_split_data = bool(args.use_dirichlet_split_data)
    args.update_feature_by_epoch = bool(args.update_feature_by_epoch)
    args.whether_initial_feature_center = bool(args.whether_initial_feature_center)
    args.teacher_repeat = bool(args.teacher_repeat)
    print("client:", args.clients_name)
    print("Whether initial_client:", args.initial_client)
    print("features_central_version_list:", args.features_central_version_list)
    if not args.use_dirichlet_split_data:
        print("client_train_num:", args.client_train_num)
        print("client_test_num:", args.client_test_num)
    print("cls_num_epochs:", args.cls_num_epochs)
    print("initial_client_ouput_feat_epoch:", args.initial_client_ouput_feat_epochs)
    print("features_ouput_layer_list:",args.features_ouput_layer_list)
    print("use_dirichlet_split_data",args.use_dirichlet_split_data)
    data = DataGenerator(args)
    client_data = data.clients[args.clients_name]
    (train_data, test_data) = client_data
    if args.dataset != 'real_data':
        if 'clients_zero_feature_index' in args:
            train_x, train_y = zip(*train_data)
            test_x, test_y = zip(*test_data) 
            train_x, test_x, data.test_x = np.array(train_x), np.array(test_x), np.array(data.test_x)
            if type(args.clients_zero_feature_index) == dict:
                print("mute feature len",len(args.clients_zero_feature_index[args.clients_name]))
                train_x[:,args.clients_zero_feature_index[args.clients_name]] = 0
                test_x[:,args.clients_zero_feature_index[args.clients_name]] = 0
                data.test_x[:,args.clients_zero_feature_index[args.clients_name]] = 0
            else:
                print("all overlap, mute feature len",len(args.clients_zero_feature_index))
                train_x[:,args.clients_zero_feature_index] = 0
                test_x[:,args.clients_zero_feature_index] = 0
                data.test_x[:,args.clients_zero_feature_index] = 0
            train_data = zip(train_x, train_y)
            test_data = zip(test_x, test_y)
        global_test_data = zip(data.test_x, data.test_y)
    else:
        args.features_ouput_layer_list = args.real_data_feature_type
        train_data = generate_dataset(train_data)
        test_data = generate_dataset(test_data)
        global_test_data = generate_dataset(data.data_test)
        train_data = train_data.map(change_dict_architecture)
        test_data = test_data.map(change_dict_architecture)
        global_test_data = global_test_data.map(change_dict_architecture)
        args.feature_space = tf.keras.utils.FeatureSpace(
            features={
                "segment": tf.keras.utils.FeatureSpace.string_categorical(output_mode='int'),
                "family": tf.keras.utils.FeatureSpace.string_categorical(output_mode='int'),
                'target': tf.keras.utils.FeatureSpace.float(),
                'tx.country': tf.keras.utils.FeatureSpace.string_categorical(output_mode='int'),
                'tx.transaction': tf.keras.utils.FeatureSpace.string_categorical(output_mode='int'),
                'tx.amount': tf.keras.utils.FeatureSpace.float(),
                'pp.p1': tf.keras.utils.FeatureSpace.float(),
                'pp.p2': tf.keras.utils.FeatureSpace.string_categorical(output_mode='int'),
            },
            output_mode='dict'
        )
        args.feature_space.adapt(train_data)
    if args.whether_initial_feature_center:
        tf.random.set_seed(args.random_seed)
        np.random.seed(args.random_seed)
        random.seed(args.random_seed)
        args.initial_feature_center = generate_initial_feature_center(args)
    model = Classifier(args)
    main(args, model, train_data, test_data, global_test_data)